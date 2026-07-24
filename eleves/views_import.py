"""
Import élèves via fichier Excel.
Colonnes attendues : Nom | Prénom | Sexe | Date naissance | Classe | Nom tuteur | Prénom tuteur | Téléphone tuteur
"""
import io
import openpyxl
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db import transaction
from eleves.models import Eleve, Tuteur, Inscription
from etablissements.models import Classe, AnneeScolaire
from accounts.models import User
import unicodedata


def _slug(name):
    """Slugifie un nom pour en faire un username."""
    n = unicodedata.normalize('NFKD', str(name)).encode('ASCII', 'ignore').decode()
    return n.lower().replace(' ', '').replace('-', '')


def _unique_username(base):
    username = base[:30]
    i = 1
    while User.objects.filter(username=username).exists():
        username = f"{base[:28]}{i}"
        i += 1
    return username


def _parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(str(val).strip()[:10], fmt).date()
        except ValueError:
            continue
    return None


@login_required
def telecharger_modele_excel(request):
    if request.user.role not in ('admin', 'super_admin', 'secretariat'):
        return redirect('dashboard')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Elèves"

    # En-têtes
    headers = [
        "Nom élève *", "Prénom élève *", "Sexe (M/F) *",
        "Date naissance (JJ/MM/AAAA)", "Classe *",
        "Nom tuteur", "Prénom tuteur", "Téléphone tuteur",
    ]
    ws.append(headers)

    # Style en-têtes
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1565C0")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Exemples
    ws.append(["DIALLO", "Aminata", "F", "15/03/2012", "6A", "DIALLO", "Moussa", "76543210"])
    ws.append(["COULIBALY", "Ibrahim", "M", "22/07/2011", "5B", "COULIBALY", "Fatoumata", "65432109"])

    # Largeurs colonnes
    for col in ws.columns:
        max_w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_w + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modele_import_eleves.xlsx"'
    wb.save(response)
    return response


@login_required
def import_eleves_excel(request):
    if request.user.role not in ('admin', 'super_admin', 'secretariat'):
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')

    etab  = request.etablissement
    annee = AnneeScolaire.objects.filter(etablissement=etab, is_active=True).first()

    if not annee:
        messages.error(request, "Aucune année scolaire active.")
        return redirect('liste_eleves')

    if request.method != "POST":
        classes = Classe.objects.filter(etablissement=etab, annee=annee).order_by('nom')
        return render(request, 'eleves/import_excel.html', {'annee': annee, 'classes': classes})

    # ── Lecture du fichier ─────────────────────────────────────────────────
    fichier = request.FILES.get('fichier_excel') or request.FILES.get('fichier')
    if not fichier:
        messages.error(request, "Aucun fichier sélectionné.")
        return redirect('import_eleves_excel')

    if not fichier.name.endswith(('.xlsx', '.xls')):
        messages.error(request, "Format non supporté — utilisez .xlsx")
        return redirect('import_eleves_excel')

    try:
        wb = openpyxl.load_workbook(fichier, data_only=True)
    except Exception as ex:
        messages.error(request, f"Fichier illisible : {ex}")
        return redirect('import_eleves_excel')

    ws = wb.active
    nb_crees = nb_ignore = 0
    erreurs = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        try:
            nom      = str(row[0] or '').strip().upper()
            prenom   = str(row[1] or '').strip().title()
            sexe_raw = str(row[2] or 'M').strip().upper()
            sexe     = 'M' if sexe_raw.startswith('M') else 'F'
            dob      = _parse_date(row[3])
            nom_cl   = str(row[4] or '').strip()
            nom_tut  = str(row[5] or '').strip().upper()
            pre_tut  = str(row[6] or '').strip().title()
            tel_tut  = str(row[7] or '').strip().replace(' ', '')

            if not nom or not prenom:
                raise ValueError("Nom et prénom obligatoires")
            if not nom_cl:
                raise ValueError("Classe obligatoire")

            # Classe
            classe = Classe.objects.filter(
                etablissement=etab, annee=annee, nom__iexact=nom_cl
            ).first()
            if not classe:
                raise ValueError(f"Classe '{nom_cl}' introuvable")

            # Doublon élève
            if Eleve.objects.filter(etablissement=etab, nom__iexact=nom, prenom__iexact=prenom).exists():
                raise ValueError(f"{nom} {prenom} déjà importé")

            with transaction.atomic():
                # Tuteur
                tuteur = None
                if nom_tut and tel_tut:
                    tuteur = Tuteur.objects.filter(
                        etablissement=etab, telephone=tel_tut
                    ).first()
                    if not tuteur:
                        u_tut = User.objects.create_user(
                            username=_unique_username(f"p{_slug(nom_tut)}{_slug(pre_tut)}"),
                            password=tel_tut[-8:] if len(tel_tut) >= 8 else 'pass1234',
                            first_name=pre_tut, last_name=nom_tut,
                            role='parent', etablissement=etab, is_active=True,
                        )
                        tuteur = Tuteur.objects.create(
                            etablissement=etab,
                            nom=nom_tut, prenom=pre_tut,
                            telephone=tel_tut, lien='pere',
                            user_compte=u_tut,
                        )

                # Matricule unique
                nb_total = Eleve.objects.filter(etablissement=etab).count()
                matricule = f"{etab.code}-{annee.date_debut.year}-{nb_total+1:04d}"
                while Eleve.objects.filter(matricule=matricule).exists():
                    nb_total += 1
                    matricule = f"{etab.code}-{annee.date_debut.year}-{nb_total:04d}"

                # Compte élève
                u_eleve = User.objects.create_user(
                    username=_unique_username(matricule.lower().replace('-', '')),
                    password=dob.strftime('%d%m%Y') if dob else 'eleve1234',
                    first_name=prenom, last_name=nom,
                    role='eleve', etablissement=etab, is_active=True,
                )

                # Élève
                eleve = Eleve.objects.create(
                    etablissement=etab, matricule=matricule,
                    nom=nom, prenom=prenom, sexe=sexe,
                    date_naissance=dob,
                    tuteur=tuteur, user_compte=u_eleve,
                    is_active=True,
                )

                # Inscription
                Inscription.objects.create(
                    eleve=eleve, classe=classe, annee=annee,
                    statut='actif', is_active=True,
                )

            nb_crees += 1

        except Exception as ex:
            nb_ignore += 1
            erreurs.append(f"Ligne {row_idx} : {ex}")

    if nb_crees:
        messages.success(request, f"✅ {nb_crees} élève(s) importé(s) avec succès.")
    if nb_ignore:
        messages.warning(request, f"⚠️ {nb_ignore} ligne(s) ignorée(s).")
        for err in erreurs[:10]:
            messages.error(request, err)

    return redirect('liste_eleves')
