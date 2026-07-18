"""
Espace super admin — Vue SaaS complète.
5 fonctionnalités :
  1. Dashboard réseau (existant, enrichi)
  2. Journal d'interventions
  3. Créer compte directeur depuis super admin
  4. Statistiques commerciales 12 mois
  5. Paramètres globaux réseau
  + Export rapport réseau Excel
"""
import datetime
import json
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q, Max
from django.http import HttpResponse
from django.utils import timezone
from .models import Etablissement, AnneeScolaire, JournalAction, ParametresReseau
from accounts.models import User
from eleves.models import Eleve
from finances.models import Paiement


def superadmin_required(fn):
    def w(request, *a, **k):
        if not request.user.is_authenticated or request.user.role != 'super_admin':
            messages.error(request, "Accès réservé au Super Administrateur.")
            return redirect('dashboard')
        return fn(request, *a, **k)
    w.__name__ = fn.__name__
    return w


def get_ip(request):
    return (request.META.get('HTTP_X_FORWARDED_FOR','').split(',')[0].strip()
            or request.META.get('REMOTE_ADDR') or None)


# ── 1. DASHBOARD RÉSEAU ───────────────────────────────────────────────────────

@login_required
@superadmin_required
def dashboard_saas(request):
    today = timezone.now().date()

    etabs = Etablissement.objects.all().order_by('nom').annotate(
        nb_eleves=Count('eleves', filter=Q(eleves__is_active=True)),
        nb_users=Count('utilisateurs', filter=Q(utilisateurs__is_active=True)),
    )

    stats = {
        'nb_etabs_actifs':      etabs.filter(is_active=True).count(),
        'nb_etabs_suspendus':   etabs.filter(is_active=False).count(),
        'nb_eleves_total':      Eleve.objects.filter(is_active=True).count(),
        'nb_users_total':       User.objects.filter(is_active=True).exclude(role='super_admin').count(),
        'recettes_reseau_mois': float(Paiement.objects.filter(
            statut='valide', date_paiement__month=today.month,
            date_paiement__year=today.year,
        ).aggregate(t=Sum('montant'))['t'] or 0),
        'recettes_reseau_annee': float(Paiement.objects.filter(
            statut='valide', date_paiement__year=today.year,
        ).aggregate(t=Sum('montant'))['t'] or 0),
    }

    semaine_debut = timezone.now() - datetime.timedelta(days=7)
    for e in etabs:
        e.recettes_mois = float(Paiement.objects.filter(
            etablissement=e, statut='valide',
            date_paiement__month=today.month, date_paiement__year=today.year,
        ).aggregate(t=Sum('montant'))['t'] or 0)
        e.connexions_semaine = User.objects.filter(
            etablissement=e, last_login__gte=semaine_debut
        ).count()
        e.sante = ('active' if e.connexions_semaine > 0 and e.is_active
                   else ('inactive' if e.is_active else 'suspendue'))

    # Graphique 12 mois recettes réseau
    chart_croissance = []
    for i in range(11, -1, -1):
        d = (today.replace(day=1) - datetime.timedelta(days=i * 28)).replace(day=1)
        recettes = float(Paiement.objects.filter(
            statut='valide', date_paiement__month=d.month, date_paiement__year=d.year,
        ).aggregate(t=Sum('montant'))['t'] or 0)
        chart_croissance.append({'mois': d.strftime('%b %Y'), 'recettes': recettes})

    etabs_inactifs_warn = [e for e in etabs if e.is_active and e.nb_eleves > 0 and e.connexions_semaine == 0]
    derniers_users = User.objects.exclude(role='super_admin').select_related('etablissement').order_by('-date_creation')[:8]
    derniers_logs  = JournalAction.objects.select_related('auteur', 'etablissement').order_by('-date')[:6]

    return render(request, 'etablissements/superadmin/dashboard_saas.html', {
        'stats': stats, 'etabs': etabs,
        'etabs_inactifs_warn': etabs_inactifs_warn,
        'chart_croissance': json.dumps(chart_croissance),
        'derniers_users': derniers_users,
        'derniers_logs': derniers_logs,
        'today': today,
        'params': ParametresReseau.get(),
    })


# ── 2. JOURNAL D'INTERVENTIONS ────────────────────────────────────────────────

@login_required
@superadmin_required
def journal_actions(request):
    """Journal complet des interventions du super admin."""
    # Filtres
    type_filtre = request.GET.get('type', '')
    etab_filtre = request.GET.get('etab', '')
    date_filtre = request.GET.get('date', '')

    qs = JournalAction.objects.select_related('auteur', 'etablissement').order_by('-date')

    if type_filtre:
        qs = qs.filter(type_action=type_filtre)
    if etab_filtre:
        qs = qs.filter(etablissement_id=etab_filtre)
    if date_filtre:
        try:
            d = datetime.datetime.strptime(date_filtre, '%Y-%m-%d').date()
            qs = qs.filter(date__date=d)
        except ValueError:
            pass

    etabs = Etablissement.objects.all().order_by('nom')

    return render(request, 'etablissements/superadmin/journal_actions.html', {
        'logs': qs[:200],
        'types': JournalAction.TYPES,
        'etabs': etabs,
        'type_filtre': type_filtre,
        'etab_filtre': etab_filtre,
        'date_filtre': date_filtre,
        'nb_total': qs.count(),
    })


# ── 3. CRÉER COMPTE DIRECTEUR ─────────────────────────────────────────────────

@login_required
@superadmin_required
def creer_directeur(request, etab_pk):
    """Créer un compte directeur pour un établissement depuis le super admin."""
    etab = get_object_or_404(Etablissement, pk=etab_pk)

    # Vérifier si un directeur existe déjà
    directeur_existant = User.objects.filter(etablissement=etab, role='admin').first()

    if request.method == 'POST':
        nom       = request.POST.get('nom', '').strip()
        prenom    = request.POST.get('prenom', '').strip()
        username  = request.POST.get('username', '').strip()
        password  = request.POST.get('password', '').strip()
        email     = request.POST.get('email', '').strip()
        telephone = request.POST.get('telephone', '').strip()

        errors = []
        if not nom or not prenom:
            errors.append("Nom et prénom obligatoires.")
        if not username:
            errors.append("Identifiant obligatoire.")
        elif User.objects.filter(username=username).exists():
            errors.append(f"L'identifiant '{username}' est déjà utilisé.")
        if len(password) < 6:
            errors.append("Le mot de passe doit faire au moins 6 caractères.")

        if errors:
            for err in errors:
                messages.error(request, err)
        else:
            user = User.objects.create_user(
                username=username, password=password,
                first_name=prenom, last_name=nom,
                email=email,
                role='admin',
                etablissement=etab,
                is_active=True,
            )
            if telephone:
                user.telephone = telephone
                user.save()

            # Journaliser
            JournalAction.log(
                request, 'creer_compte',
                cible=f"{nom} {prenom} ({username})",
                detail=f"Compte directeur créé pour {etab.nom}",
                etablissement=etab,
            )

            messages.success(request, f"✅ Compte directeur '{username}' créé pour {etab.nom}.")
            return redirect('gestion_comptes_etab', pk=etab_pk)

    return render(request, 'etablissements/superadmin/creer_directeur.html', {
        'etab': etab,
        'directeur_existant': directeur_existant,
    })


# ── 4. STATISTIQUES COMMERCIALES 12 MOIS ─────────────────────────────────────

@login_required
@superadmin_required
def stats_commerciales(request):
    """Statistiques de croissance SaaS : recettes par établissement sur 12 mois."""
    today = timezone.now().date()
    annee = int(request.GET.get('annee', today.year))

    MOIS_FR = ['','Janv','Févr','Mars','Avr','Mai','Juin',
               'Juil','Août','Sept','Oct','Nov','Déc']

    etabs = Etablissement.objects.all().order_by('nom')

    # Recettes par établissement × mois (tableau croisé)
    tableau = []
    totaux_mois = [0.0] * 12
    grand_total = 0.0

    for e in etabs:
        ligne = {'etab': e, 'mois': [], 'total': 0.0}
        for m in range(1, 13):
            t = float(Paiement.objects.filter(
                etablissement=e, statut='valide',
                date_paiement__month=m, date_paiement__year=annee,
            ).aggregate(t=Sum('montant'))['t'] or 0)
            ligne['mois'].append(t)
            ligne['total'] += t
            totaux_mois[m - 1] += t
            grand_total += t
        tableau.append(ligne)

    # Graphique comparatif : une courbe par établissement
    chart_etabs = []
    for e in etabs:
        serie = []
        for m in range(1, 13):
            t = float(Paiement.objects.filter(
                etablissement=e, statut='valide',
                date_paiement__month=m, date_paiement__year=annee,
            ).aggregate(t=Sum('montant'))['t'] or 0)
            serie.append(t)
        chart_etabs.append({'nom': e.nom, 'data': serie})

    # Évolution recettes réseau mois par mois
    chart_reseau = [{'mois': MOIS_FR[m], 'total': totaux_mois[m-1]} for m in range(1, 13)]

    # Top établissements par recettes
    top_etabs = sorted(tableau, key=lambda x: x['total'], reverse=True)[:5]

    return render(request, 'etablissements/superadmin/stats_commerciales.html', {
        'tableau': tableau,
        'totaux_mois': totaux_mois,
        'grand_total': grand_total,
        'mois_labels': [MOIS_FR[m] for m in range(1, 13)],
        'chart_etabs': json.dumps(chart_etabs),
        'chart_reseau': json.dumps(chart_reseau),
        'top_etabs': top_etabs,
        'annee': annee,
        'annees': range(today.year - 2, today.year + 1),
        'today': today,
    })


# ── 5. PARAMÈTRES GLOBAUX RÉSEAU ─────────────────────────────────────────────

@login_required
@superadmin_required
def parametres_reseau(request):
    """Paramètres globaux de la plateforme SmartSchool."""
    params = ParametresReseau.get()

    if request.method == 'POST':
        params.nom_plateforme     = request.POST.get('nom_plateforme', params.nom_plateforme).strip()
        params.slogan             = request.POST.get('slogan', '').strip()
        params.email_support      = request.POST.get('email_support', '').strip()
        params.telephone_support  = request.POST.get('telephone_support', '').strip()
        params.site_web           = request.POST.get('site_web', '').strip()
        params.couleur_primaire   = request.POST.get('couleur_primaire', '#1565C0')
        params.couleur_secondaire = request.POST.get('couleur_secondaire', '#0D47A1')
        params.mentions_legales   = request.POST.get('mentions_legales', '').strip()
        params.message_accueil    = request.POST.get('message_accueil', '').strip()
        if request.FILES.get('logo_plateforme'):
            params.logo_plateforme = request.FILES['logo_plateforme']
        params.save()

        JournalAction.log(
            request, 'autre',
            cible='Paramètres réseau',
            detail='Modification des paramètres globaux de la plateforme',
        )

        messages.success(request, "✅ Paramètres sauvegardés.")
        return redirect('parametres_reseau')

    from django.conf import settings
    import django
    infos_systeme = {
        'Version Django': django.__version__,
        'Version SmartSchool': 'v3.0',
        'Établissements': Etablissement.objects.count(),
        'Élèves total': Eleve.objects.filter(is_active=True).count(),
        'Utilisateurs total': User.objects.filter(is_active=True).exclude(role='super_admin').count(),
        'Debug': 'Oui' if settings.DEBUG else 'Non',
    }
    return render(request, 'etablissements/superadmin/parametres_reseau.html', {
        'params': params,
        'infos_systeme': infos_systeme,
    })


# ── EXPORT RAPPORT RÉSEAU EXCEL ───────────────────────────────────────────────

@login_required
@superadmin_required
def export_rapport_reseau(request):
    """Export Excel de l'état complet du réseau SmartSchool."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        messages.error(request, "openpyxl non installé.")
        return redirect('dashboard_saas')

    today = timezone.now().date()
    annee = int(request.GET.get('annee', today.year))

    wb = openpyxl.Workbook()
    BLEU = "1565C0"; BLEU_CLAIR = "E3F2FD"; VERT = "2E7D32"; ROUGE = "C62828"
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor=BLEU)
    halign = Alignment(horizontal="center", vertical="center")
    thin = Side(style='thin', color="CCCCCC")
    border = lambda: Border(left=thin, right=thin, top=thin, bottom=thin)

    # Feuille 1 : Vue réseau
    ws1 = wb.active
    ws1.title = "Vue réseau"
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 10
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 16
    ws1.column_dimensions['F'].width = 14

    ws1.merge_cells('A1:F1')
    ws1['A1'] = f"RAPPORT RÉSEAU SMARTSCHOOL — {annee}"
    ws1['A1'].font = Font(bold=True, size=14, color=BLEU)
    ws1['A1'].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells('A2:F2')
    ws1['A2'] = f"Exporté le {today.strftime('%d/%m/%Y')} | {Etablissement.objects.count()} établissement(s)"
    ws1['A2'].font = Font(italic=True, size=10, color="666666")
    ws1['A2'].alignment = Alignment(horizontal="center")

    headers = ['Établissement', 'Élèves', 'Staff', 'Recettes mois', 'Recettes année', 'Statut']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = halign
        cell.border = border()
    ws1.row_dimensions[4].height = 22

    semaine_debut = timezone.now() - datetime.timedelta(days=7)
    etabs = Etablissement.objects.all().order_by('nom').annotate(
        nb_eleves=Count('eleves', filter=Q(eleves__is_active=True)),
        nb_users=Count('utilisateurs', filter=Q(utilisateurs__is_active=True)),
    )

    row = 5
    for i, e in enumerate(etabs):
        recettes_mois = float(Paiement.objects.filter(
            etablissement=e, statut='valide',
            date_paiement__month=today.month, date_paiement__year=today.year,
        ).aggregate(t=Sum('montant'))['t'] or 0)
        recettes_annee = float(Paiement.objects.filter(
            etablissement=e, statut='valide', date_paiement__year=annee,
        ).aggregate(t=Sum('montant'))['t'] or 0)
        connexions = User.objects.filter(etablissement=e, last_login__gte=semaine_debut).count()
        statut = "Actif" if e.is_active and connexions > 0 else ("Dormant" if e.is_active else "Suspendu")

        vals = [e.nom, e.nb_eleves, e.nb_users, recettes_mois, recettes_annee, statut]
        alt = PatternFill("solid", fgColor="F5F5F5") if i % 2 else None
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.border = border()
            if alt: cell.fill = alt
            if col in (4, 5):
                cell.number_format = '#,##0'
                cell.font = Font(color=VERT if val > 0 else "999999")
            if col == 6:
                cell.font = Font(color=(VERT if statut == "Actif" else (ROUGE if statut == "Suspendu" else "F57F17")), bold=True)
        row += 1

    # Ligne total
    ws1.cell(row=row, column=1, value="TOTAL RÉSEAU").font = Font(bold=True, size=11)
    t_mois = sum(float(Paiement.objects.filter(statut='valide', date_paiement__month=today.month, date_paiement__year=today.year).aggregate(t=Sum('montant'))['t'] or 0) for _ in [1])
    t_ann  = sum(float(Paiement.objects.filter(statut='valide', date_paiement__year=annee).aggregate(t=Sum('montant'))['t'] or 0) for _ in [1])
    for col, val in [(4, t_mois), (5, t_ann)]:
        cell = ws1.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True, size=12, color=BLEU)
        cell.number_format = '#,##0'
        cell.fill = PatternFill("solid", fgColor=BLEU_CLAIR)
        cell.border = border()

    # Feuille 2 : Journal actions
    ws2 = wb.create_sheet("Journal actions")
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 24
    ws2.column_dimensions['E'].width = 30

    for col, h in enumerate(['Date', 'Type action', 'Établissement', 'Cible', 'Détail'], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = halign; cell.border = border()

    for r, log in enumerate(JournalAction.objects.select_related('etablissement').order_by('-date')[:100], 2):
        ws2.cell(row=r, column=1, value=log.date.strftime('%d/%m/%Y %H:%M'))
        ws2.cell(row=r, column=2, value=log.get_type_action_display())
        ws2.cell(row=r, column=3, value=log.etablissement.nom if log.etablissement else '—')
        ws2.cell(row=r, column=4, value=log.cible)
        ws2.cell(row=r, column=5, value=log.detail)
        for col in range(1, 6):
            ws2.cell(row=r, column=col).border = border()

    # Journaliser l'export
    JournalAction.log(request, 'export', cible=f'Rapport réseau {annee}', detail='Export Excel rapport réseau complet')

    nom = f"smartschool_rapport_reseau_{annee}_{today.strftime('%d%m%Y')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    wb.save(response)
    return response


# ── Journaliser dans views_superadmin (patch) ─────────────────────────────────

def log_action(request, type_action, cible='', detail='', etablissement=None):
    """Raccourci utilisé par views_superadmin pour journaliser."""
    JournalAction.log(request, type_action, cible=cible, detail=detail, etablissement=etablissement)
