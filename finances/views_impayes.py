"""
Frais impayés : liste par classe + export Excel + génération de notifications.
"""
import datetime
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from django.utils import timezone
from accounts.permissions import permission_required
from .models import Paiement, TypeFrais
from eleves.models import Eleve, Inscription
from etablissements.models import AnneeScolaire, Classe
from core.cycle_filter import get_eleves_actifs, get_classes_actives


def _get_impayes(etab, annee, classe=None, type_frais=None, mois=None, annee_val=None):
    """
    Retourne la liste des élèves avec leurs frais impayés.
    Un élève est 'impayé' sur un type de frais s'il n'a aucun paiement valide
    pour ce type sur la période demandée.
    """
    today = timezone.now().date()
    mois = mois or today.month
    annee_val = annee_val or today.year

    # Élèves actifs (filtrés par classe si demandé)
    if classe:
        inscriptions = Inscription.objects.filter(
            classe=classe, annee=annee, is_active=True
        ).select_related('eleve', 'classe')
        eleves_qs = [i.eleve for i in inscriptions]
    else:
        eleves_qs = list(get_eleves_actifs(etab))

    if not eleves_qs:
        return []

    # Types de frais à vérifier
    types = TypeFrais.objects.filter(etablissement=etab)
    if type_frais:
        types = types.filter(pk=type_frais.pk)
    if annee:
        types = types.filter(Q(annee=annee) | Q(annee__isnull=True))

    # IDs d'élèves ayant payé ce mois pour chaque type
    eleves_ids = [e.pk for e in eleves_qs]

    # Construire la liste des impayés
    impayes = []
    for eleve in eleves_qs:
        # Paiements valides de cet élève ce mois
        paiements_mois = Paiement.objects.filter(
            eleve=eleve, etablissement=etab, statut='valide',
            date_paiement__month=mois, date_paiement__year=annee_val,
        ).values_list('type_frais_id', flat=True)

        # Inscription active pour avoir la classe
        insc = Inscription.objects.filter(
            eleve=eleve, annee=annee, is_active=True
        ).select_related('classe').first()

        # Pour chaque type de frais non payé
        frais_non_payes = []
        for tf in types:
            if tf.pk not in paiements_mois:
                # Vérifier aussi les paiements annuels (periodicite='unique')
                if tf.periodicite == 'unique':
                    paye_annee = Paiement.objects.filter(
                        eleve=eleve, etablissement=etab, statut='valide',
                        type_frais=tf,
                        date_paiement__year=annee_val,
                    ).exists()
                    if paye_annee:
                        continue
                frais_non_payes.append(tf)

        if frais_non_payes:
            # Dernier paiement de l'élève (toutes catégories)
            dernier = Paiement.objects.filter(
                eleve=eleve, etablissement=etab, statut='valide'
            ).order_by('-date_paiement').first()

            impayes.append({
                'eleve': eleve,
                'classe': insc.classe if insc else None,
                'frais_non_payes': frais_non_payes,
                'nb_frais': len(frais_non_payes),
                'montant_du': sum(float(tf.montant_defaut) for tf in frais_non_payes),
                'dernier_paiement': dernier.date_paiement if dernier else None,
            })

    # Trier par classe puis par nom
    impayes.sort(key=lambda x: (
        x['classe'].nom if x['classe'] else 'ZZZ',
        x['eleve'].nom,
        x['eleve'].prenom,
    ))
    return impayes


@login_required
@permission_required('finances')
def liste_impayes(request):
    """Liste des frais impayés avec filtres par classe, type de frais, mois."""
    etab = request.etablissement
    annee = AnneeScolaire.objects.filter(etablissement=etab, is_active=True).first()
    today = timezone.now().date()

    # Filtres
    classe_id  = request.GET.get('classe', '')
    type_id    = request.GET.get('type_frais', '')
    mois       = int(request.GET.get('mois', today.month))
    annee_val  = int(request.GET.get('annee', today.year))

    classe_sel = Classe.objects.filter(pk=classe_id, etablissement=etab).first() if classe_id else None
    type_sel   = TypeFrais.objects.filter(pk=type_id, etablissement=etab).first() if type_id else None

    impayes = _get_impayes(etab, annee, classe=classe_sel, type_frais=type_sel,
                           mois=mois, annee_val=annee_val)

    # Regrouper par classe
    par_classe = {}
    for imp in impayes:
        cl_nom = imp['classe'].nom if imp['classe'] else 'Sans classe'
        if cl_nom not in par_classe:
            par_classe[cl_nom] = {'classe': imp['classe'], 'eleves': [], 'total_du': 0}
        par_classe[cl_nom]['eleves'].append(imp)
        par_classe[cl_nom]['total_du'] += imp['montant_du']

    # Stats globales
    total_eleves_impayes = len(impayes)
    total_montant_du = sum(i['montant_du'] for i in impayes)

    MOIS_FR = ['','Janvier','Février','Mars','Avril','Mai','Juin',
               'Juillet','Août','Septembre','Octobre','Novembre','Décembre']

    classes   = get_classes_actives(etab, annee)
    types_frais = TypeFrais.objects.filter(
        etablissement=etab
    ).filter(Q(annee=annee) | Q(annee__isnull=True)) if annee else TypeFrais.objects.filter(etablissement=etab)

    return render(request, 'finances/liste_impayes.html', {
        'impayes': impayes,
        'par_classe': par_classe,
        'total_eleves_impayes': total_eleves_impayes,
        'total_montant_du': total_montant_du,
        'classes': classes,
        'types_frais': types_frais,
        'classe_sel': classe_sel,
        'type_sel': type_sel,
        'mois': mois,
        'annee_val': annee_val,
        'mois_fr': MOIS_FR[mois],
        'mois_choices': [(i, MOIS_FR[i]) for i in range(1, 13)],
        'annees': range(today.year - 1, today.year + 1),
        'annee': annee,
        'today': today,
    })


@login_required
@permission_required('finances')
def export_impayes_excel(request):
    """Export Excel des impayés filtré par classe / type de frais / mois."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from django.contrib import messages
        messages.error(request, "openpyxl non installé.")
        return render(request, 'finances/liste_impayes.html', {})

    etab = request.etablissement
    annee = AnneeScolaire.objects.filter(etablissement=etab, is_active=True).first()
    today = timezone.now().date()

    classe_id = request.GET.get('classe', '')
    type_id   = request.GET.get('type_frais', '')
    mois      = int(request.GET.get('mois', today.month))
    annee_val = int(request.GET.get('annee', today.year))

    classe_sel = Classe.objects.filter(pk=classe_id, etablissement=etab).first() if classe_id else None
    type_sel   = TypeFrais.objects.filter(pk=type_id, etablissement=etab).first() if type_id else None

    impayes = _get_impayes(etab, annee, classe=classe_sel, type_frais=type_sel,
                           mois=mois, annee_val=annee_val)

    MOIS_FR = ['','Janvier','Février','Mars','Avril','Mai','Juin',
               'Juillet','Août','Septembre','Octobre','Novembre','Décembre']

    # ── Styles ────────────────────────────────────────────────────────────────
    ROUGE = "C62828"; ROUGE_CLAIR = "FFEBEE"; BLEU = "1565C0"; GRIS = "F5F5F5"
    hf    = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor=ROUGE)
    halign= Alignment(horizontal="center", vertical="center")
    thin  = Side(style='thin', color="CCCCCC")
    def brd(): return Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Impayés {MOIS_FR[mois]}"

    # Largeurs colonnes
    for col, w in zip('ABCDEFG', [8, 28, 16, 24, 16, 22, 18]):
        ws.column_dimensions[col].width = w

    # Titre
    ws.merge_cells('A1:G1')
    ws['A1'] = f"FRAIS IMPAYÉS — {MOIS_FR[mois].upper()} {annee_val}"
    ws['A1'].font = Font(bold=True, size=14, color=ROUGE)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:G2')
    info = f"{etab.nom}"
    if classe_sel: info += f" | Classe : {classe_sel.nom}"
    if type_sel:   info += f" | Frais : {type_sel.nom}"
    info += f" | Exporté le {today.strftime('%d/%m/%Y')}"
    ws['A2'] = info
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws['A2'].alignment = Alignment(horizontal="center")

    # Résumé
    ws.merge_cells('A3:G3')
    ws['A3'] = f"Total élèves en retard : {len(impayes)} | Montant dû estimé : {sum(i['montant_du'] for i in impayes):,.0f} FCFA"
    ws['A3'].font = Font(bold=True, size=11, color=ROUGE)
    ws['A3'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 20

    # En-têtes
    headers = ['#', 'Élève', 'Classe', 'Frais non payé(s)', 'Montant dû (FCFA)', 'Contact tuteur', 'Dernier paiement']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = halign; cell.border = brd()
    ws.row_dimensions[5].height = 22

    # Données
    row = 6
    current_classe = None
    for i, imp in enumerate(impayes, 1):
        # Séparateur de classe
        if imp['classe'] and (current_classe != imp['classe'].nom):
            current_classe = imp['classe'].nom
            ws.merge_cells(f'A{row}:G{row}')
            ws.cell(row=row, column=1, value=f"📚 {current_classe}").font = Font(bold=True, color=BLEU, size=11)
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="E3F2FD")
            for c in range(1, 8):
                ws.cell(row=row, column=c).border = brd()
            row += 1

        # Tuteur
        try:
            tut = imp['eleve'].tuteur
            tut_info = f"{tut.nom} {tut.prenom}" if tut else "—"
            tut_tel  = tut.telephone if tut else ""
            contact  = f"{tut_info} | {tut_tel}" if tut_tel else tut_info
        except Exception:
            contact = "—"

        frais_noms = ", ".join(tf.nom for tf in imp['frais_non_payes'])
        alt = PatternFill("solid", fgColor=GRIS) if i % 2 == 0 else None

        vals = [i, imp['eleve'].nom_complet, imp['classe'].nom if imp['classe'] else '—',
                frais_noms, imp['montant_du'], contact,
                imp['dernier_paiement'].strftime('%d/%m/%Y') if imp['dernier_paiement'] else 'Jamais']

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = brd()
            if alt: cell.fill = alt
            if col == 5:
                cell.number_format = '#,##0'
                cell.font = Font(bold=True, color=ROUGE)
        ws.row_dimensions[row].height = 18
        row += 1

    # Total final
    total = sum(i['montant_du'] for i in impayes)
    ws.cell(row=row, column=4, value="TOTAL DÛ").font = Font(bold=True)
    cell_t = ws.cell(row=row, column=5, value=total)
    cell_t.font = Font(bold=True, size=12, color=ROUGE)
    cell_t.number_format = '#,##0'
    cell_t.fill = PatternFill("solid", fgColor=ROUGE_CLAIR)
    for c in range(1, 8):
        ws.cell(row=row, column=c).border = brd()

    ws.freeze_panes = 'A6'

    # Journaliser
    from etablissements.models import JournalAction
    JournalAction.log(request, 'export',
                      cible=f'Impayés {MOIS_FR[mois]} {annee_val}',
                      detail=f'{len(impayes)} élèves en retard')

    nom = f"impayes_{etab.code}_{mois:02d}{annee_val}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    wb.save(response)
    return response


def creer_notifications_impayes(etab, annee):
    """
    Génère des notifications 'paiement' pour le comptable
    sur les élèves en retard de paiement ce mois.
    Appelée depuis :
      - Le dashboard comptable (lazy, une fois par jour)
      - Un management command cron
    """
    from core.models import Notification
    from accounts.models import User

    today = timezone.now().date()
    impayes = _get_impayes(etab, annee, mois=today.month, annee_val=today.year)
    if not impayes:
        return 0

    # Trouver les comptables de l'établissement
    comptables = User.objects.filter(etablissement=etab, role='comptable', is_active=True)
    if not comptables.exists():
        # Notifier aussi le directeur
        comptables = User.objects.filter(etablissement=etab, role='admin', is_active=True)

    if not comptables.exists():
        return 0

    # Ne pas créer de doublon pour aujourd'hui
    deja = Notification.objects.filter(
        destinataire__in=comptables,
        type_notif=Notification.TYPE_PAIEMENT,
        date_creation__date=today,
        titre__startswith='Frais impayés',
    ).exists()

    if deja:
        return 0

    nb_created = 0
    for comptable in comptables:
        nb = len(impayes)
        total = sum(i['montant_du'] for i in impayes)
        # Détail par classe
        classes_detail = {}
        for imp in impayes:
            cl = imp['classe'].nom if imp['classe'] else 'Sans classe'
            classes_detail[cl] = classes_detail.get(cl, 0) + 1
        detail = ' | '.join(f"{cl}: {n}" for cl, n in list(classes_detail.items())[:5])

        Notification.objects.create(
            destinataire=comptable,
            type_notif=Notification.TYPE_PAIEMENT,
            titre=f'Frais impayés — {nb} élève(s) ce mois',
            message=(
                f"{nb} élève(s) n'ont pas encore réglé leurs frais pour ce mois.\n"
                f"Montant total estimé : {total:,.0f} FCFA\n"
                f"Répartition : {detail}"
            ),
            lien='/finances/impayes/',
        )
        nb_created += 1

    return nb_created
