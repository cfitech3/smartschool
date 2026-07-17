"""
Export Excel et PDF des données financières.
"""
import io
import datetime
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.utils import timezone
from accounts.permissions import permission_required
from .models import Paiement, TypeFrais
from eleves.models import Eleve, Inscription
from core.cycle_filter import get_eleves_actifs


@login_required
@permission_required('finances')
def export_excel_finances(request):
    """Export Excel des paiements avec filtres."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, "openpyxl non installé. Lancez : pip install openpyxl")
        return redirect('rapport_financier')

    etab = request.etablissement
    today = timezone.now().date()

    # Filtres
    mois = int(request.GET.get('mois', today.month))
    annee_val = int(request.GET.get('annee', today.year))
    type_export = request.GET.get('type', 'paiements')  # paiements | retards | bilan

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    BLEU = "1565C0"
    BLEU_CLAIR = "E3F2FD"
    VERT = "2E7D32"
    ROUGE = "C62828"
    GRIS = "F5F5F5"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=BLEU)
    header_align = Alignment(horizontal="center", vertical="center")
    subheader_fill = PatternFill("solid", fgColor=BLEU_CLAIR)
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, row, cols):
        for col in range(1, cols+1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

    def style_row(ws, row, cols, alt=False):
        fill = PatternFill("solid", fgColor=GRIS) if alt else None
        for col in range(1, cols+1):
            cell = ws.cell(row=row, column=col)
            if fill: cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    MOIS_FR = ['', 'Janvier','Février','Mars','Avril','Mai','Juin',
               'Juillet','Août','Septembre','Octobre','Novembre','Décembre']

    if type_export == 'paiements':
        ws = wb.active
        ws.title = f"Paiements {MOIS_FR[mois]}"
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 14

        # Titre
        ws.merge_cells('A1:H1')
        ws['A1'] = f"ÉTAT DES PAIEMENTS — {MOIS_FR[mois].upper()} {annee_val}"
        ws['A1'].font = Font(bold=True, size=14, color=BLEU)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells('A2:H2')
        ws['A2'] = f"{etab.nom} | Exporté le {today.strftime('%d/%m/%Y')}"
        ws['A2'].font = Font(italic=True, size=10, color="666666")
        ws['A2'].alignment = Alignment(horizontal="center")

        ws.row_dimensions[3].height = 5  # espace

        # En-têtes
        entetes = ['#', 'Élève', 'Classe', 'Type de frais', 'Montant (FCFA)', 'Mode', 'Date', 'Réf.']
        for col, titre in enumerate(entetes, 1):
            ws.cell(row=4, column=col, value=titre)
        style_header(ws, 4, len(entetes))
        ws.row_dimensions[4].height = 22

        # Données
        paiements = Paiement.objects.filter(
            etablissement=etab, statut='valide',
            date_paiement__month=mois, date_paiement__year=annee_val
        ).select_related('eleve', 'type_frais').prefetch_related('eleve__inscriptions__classe').order_by('-date_paiement')

        total = 0
        row = 5
        for i, p in enumerate(paiements, 1):
            insc = p.eleve.inscriptions.filter(is_active=True).first()
            classe_nom = insc.classe.nom if insc else '—'
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=p.eleve.nom_complet)
            ws.cell(row=row, column=3, value=classe_nom)
            ws.cell(row=row, column=4, value=p.type_frais.nom if p.type_frais else '—')
            montant_cell = ws.cell(row=row, column=5, value=float(p.montant))
            montant_cell.number_format = '#,##0'
            montant_cell.font = Font(bold=True, color=VERT)
            ws.cell(row=row, column=6, value=p.get_mode_paiement_display())
            ws.cell(row=row, column=7, value=p.date_paiement.strftime('%d/%m/%Y %H:%M') if p.date_paiement else '—')
            ws.cell(row=row, column=8, value=p.reference or '—')
            style_row(ws, row, 8, alt=(i % 2 == 0))
            ws.row_dimensions[row].height = 18
            total += float(p.montant)
            row += 1

        # Total
        ws.cell(row=row, column=4, value="TOTAL").font = Font(bold=True)
        total_cell = ws.cell(row=row, column=5, value=total)
        total_cell.font = Font(bold=True, size=12, color=VERT)
        total_cell.number_format = '#,##0'
        total_cell.fill = PatternFill("solid", fgColor="E8F5E9")
        for c in range(1, 9):
            ws.cell(row=row, column=c).border = border

        ws.freeze_panes = 'A5'

    elif type_export == 'retards':
        ws = wb.active
        ws.title = "Retards de paiement"
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 16

        ws.merge_cells('A1:F1')
        ws['A1'] = f"ÉLÈVES EN RETARD DE PAIEMENT — {MOIS_FR[mois].upper()} {annee_val}"
        ws['A1'].font = Font(bold=True, size=14, color="C62828")
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells('A2:F2')
        ws['A2'] = f"{etab.nom} | Exporté le {today.strftime('%d/%m/%Y')}"
        ws['A2'].font = Font(italic=True, size=10, color="666666")
        ws['A2'].alignment = Alignment(horizontal="center")

        entetes = ['#', 'Élève', 'Classe', 'Contact tuteur', 'Téléphone', 'Dernier paiement']
        for col, titre in enumerate(entetes, 1):
            ws.cell(row=4, column=col, value=titre)
        style_header(ws, 4, 6)
        ws.row_dimensions[4].height = 22

        payes_ids = Paiement.objects.filter(
            etablissement=etab, statut='valide',
            date_paiement__month=mois, date_paiement__year=annee_val
        ).values_list('eleve_id', flat=True).distinct()

        eleves_retard = get_eleves_actifs(etab).exclude(pk__in=payes_ids).prefetch_related('tuteur', 'inscriptions__classe').order_by('nom', 'prenom')

        row = 5
        for i, eleve in enumerate(eleves_retard, 1):
            insc = eleve.inscriptions.filter(is_active=True).first()
            classe_nom = insc.classe.nom if insc else '—'
            tuteur = eleve.tuteur if hasattr(eleve, 'tuteur') and eleve.tuteur else None
            tuteur_nom = f'{tuteur.nom} {tuteur.prenom}' if tuteur else '—'
            tuteur_tel = tuteur.telephone if tuteur else '—'
            dernier = Paiement.objects.filter(
                eleve=eleve, etablissement=etab, statut='valide'
            ).order_by('-date_paiement').first()
            dernier_str = dernier.date_paiement.strftime('%d/%m/%Y') if dernier else 'Jamais'

            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=eleve.nom_complet).font = Font(bold=True)
            ws.cell(row=row, column=3, value=classe_nom)
            ws.cell(row=row, column=4, value=tuteur_nom)
            ws.cell(row=row, column=5, value=tuteur_tel)
            retard_cell = ws.cell(row=row, column=6, value=dernier_str)
            retard_cell.font = Font(color=ROUGE)
            style_row(ws, row, 6, alt=(i % 2 == 0))
            ws.row_dimensions[row].height = 18
            row += 1

        ws.cell(row=row, column=1, value=f"Total : {eleves_retard.count()} élève(s) en retard")
        ws.cell(row=row, column=1).font = Font(bold=True, color=ROUGE)
        ws.freeze_panes = 'A5'

    elif type_export == 'bilan':
        ws = wb.active
        ws.title = f"Bilan {annee_val}"
        ws.column_dimensions['A'].width = 20
        for c in 'BCDEFGHIJKL':
            ws.column_dimensions[c].width = 12

        ws.merge_cells('A1:M1')
        ws['A1'] = f"BILAN FINANCIER ANNUEL {annee_val} — {etab.nom}"
        ws['A1'].font = Font(bold=True, size=14, color=BLEU)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30

        # En-tête mois
        ws.cell(row=3, column=1, value="Type de frais").font = Font(bold=True)
        for m_idx, m_nom in enumerate(MOIS_FR[1:], 2):
            cell = ws.cell(row=3, column=m_idx, value=m_nom[:4])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.cell(row=3, column=14, value="TOTAL").font = header_font
        ws.cell(row=3, column=14).fill = header_fill
        ws.cell(row=3, column=14).alignment = header_align

        types = TypeFrais.objects.filter(etablissement=etab)
        row = 4
        totaux_mois = [0] * 12
        for tf in types:
            ws.cell(row=row, column=1, value=tf.nom).font = Font(bold=True)
            total_ligne = 0
            for m_idx in range(1, 13):
                montant = float(Paiement.objects.filter(
                    etablissement=etab, type_frais=tf, statut='valide',
                    date_paiement__month=m_idx, date_paiement__year=annee_val
                ).aggregate(t=Sum('montant'))['t'] or 0)
                cell = ws.cell(row=row, column=m_idx+1, value=montant if montant else '')
                if montant:
                    cell.number_format = '#,##0'
                    cell.font = Font(color=VERT)
                totaux_mois[m_idx-1] += montant
                total_ligne += montant
            total_cell = ws.cell(row=row, column=14, value=total_ligne)
            total_cell.number_format = '#,##0'
            total_cell.font = Font(bold=True, color=VERT)
            style_row(ws, row, 14, alt=(row % 2 == 0))
            row += 1

        # Ligne totaux
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, size=11)
        grand_total = 0
        for m_idx, t in enumerate(totaux_mois, 2):
            cell = ws.cell(row=row, column=m_idx, value=t if t else '')
            cell.font = Font(bold=True, color=BLEU)
            cell.number_format = '#,##0'
            cell.fill = PatternFill("solid", fgColor=BLEU_CLAIR)
            grand_total += t
        cell_gt = ws.cell(row=row, column=14, value=grand_total)
        cell_gt.font = Font(bold=True, size=12, color=BLEU)
        cell_gt.number_format = '#,##0'
        cell_gt.fill = PatternFill("solid", fgColor=BLEU_CLAIR)
        ws.freeze_panes = 'B4'

    # Réponse HTTP
    nom_fichier = f"smartschool_{type_export}_{etab.code}_{mois:02d}{annee_val}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
    wb.save(response)
    return response
