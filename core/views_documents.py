from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q, Sum
from eleves.models import Eleve
from finances.models import Paiement
from etablissements.models import AnneeScolaire, ModeleDocument
from core.cycle_filter import get_cycles_actifs_ids, get_classes_actives, get_eleves_actifs, get_inscriptions_actives

def req(fn):
    def w(request,*a,**k):
        if not request.etablissement: return redirect("dashboard")
        return fn(request,*a,**k)
    w.__name__=fn.__name__; return w

def get_modele(etab, type_doc):
    """Récupère le modèle actif. Gère les alias entre URL et type en base."""
    # Mapping URL param -> type en base
    alias = {
        'certificat_scolarite': 'certificat',
        'attestation_frequentation': 'attestation',
        'recu_paiement_annuel': 'recu',
        'carte_scolaire': 'carte_scolaire',
        'releve_notes': 'releve_notes',
        'bulletin': 'bulletin',
    }
    type_reel = alias.get(type_doc, type_doc)
    return ModeleDocument.objects.filter(
        etablissement=etab, type_document=type_reel, is_actif=True
    ).first()

@login_required
@req
def recherche_globale(request):
    q=request.GET.get("q","").strip(); etab=request.etablissement
    resultats={"eleves":[],"paiements":[],"enseignants":[]}
    if q and len(q)>=2:
        from accounts.models import User
        resultats["eleves"]=get_eleves_actifs(etab).filter(Q(nom__icontains=q)|Q(prenom__icontains=q)|Q(matricule__icontains=q))[:10]
        resultats["paiements"]=Paiement.objects.filter(etablissement=etab).filter(Q(eleve__nom__icontains=q)|Q(eleve__prenom__icontains=q)|Q(reference__icontains=q)).select_related("eleve","type_frais")[:5]
        resultats["enseignants"]=User.objects.filter(etablissement=etab,role="enseignant").filter(Q(first_name__icontains=q)|Q(last_name__icontains=q))[:5]
    return render(request,"core/recherche.html",{"q":q,"resultats":resultats,"total":sum(len(v) for v in resultats.values())})

@login_required
@req
def liste_documents(request):
    etab=request.etablissement; annee=AnneeScolaire.objects.filter(etablissement=etab,is_active=True).first()
    q=request.GET.get("q","")
    eleves=get_eleves_actifs(etab).order_by("nom","prenom")
    if q: eleves=eleves.filter(Q(nom__icontains=q)|Q(prenom__icontains=q)|Q(matricule__icontains=q))
    modeles_actifs={m.type_document:m for m in ModeleDocument.objects.filter(etablissement=etab,is_actif=True)}
    return render(request,"core/documents.html",{"eleves":eleves,"annee":annee,"q":q,"modeles_actifs":modeles_actifs})

@login_required
@req
def generer_document(request, eleve_pk, type_doc):
    etab=request.etablissement
    eleve=get_object_or_404(Eleve,pk=eleve_pk,etablissement=etab)
    annee=AnneeScolaire.objects.filter(etablissement=etab,is_active=True).first()
    inscription=eleve.get_inscription_active(); today=timezone.now().date()
    # Récupérer le bon modèle selon le type
    modele=get_modele(etab, type_doc)
    # Remplacement des balises si un contenu personnalisé existe
    contenu_rendu = None
    if modele and modele.contenu_personnalise:
        contenu_rendu = modele.contenu_personnalise
        contenu_rendu = contenu_rendu.replace('[NOM_ELEVE]', f"{eleve.nom} {eleve.prenom}")
        contenu_rendu = contenu_rendu.replace('[CLASSE]', inscription.classe.nom if inscription else '')
        contenu_rendu = contenu_rendu.replace('[MATRICULE]', eleve.matricule or '')
        date_n = eleve.date_naissance.strftime('%d/%m/%Y') if eleve.date_naissance else ''
        contenu_rendu = contenu_rendu.replace('[DATE_NAISSANCE]', date_n)
        contenu_rendu = contenu_rendu.replace('[LIEU_NAISSANCE]', eleve.lieu_naissance or '')
        contenu_rendu = contenu_rendu.replace('[ANNEE_SCOLAIRE]', annee.libelle if annee else '')
        contenu_rendu = contenu_rendu.replace('[NOM_ECOLE]', etab.nom)
    
    ctx={"eleve":eleve,"etab":etab,"annee":annee,"inscription":inscription,
         "today":today,"modele":modele,
         "ref":f"{type_doc[:4].upper()}-{eleve.matricule[-4:]}-{today.year}",
         "contenu_personnalise_rendu": contenu_rendu}

    if type_doc=="certificat_scolarite":
        return render(request,"core/docs/certificat_scolarite.html",ctx)
    elif type_doc=="attestation_frequentation":
        return render(request,"core/docs/attestation_frequentation.html",ctx)
    elif type_doc=="recu_paiement_annuel":
        paiements=eleve.paiements.filter(annee=annee,statut="valide").select_related("type_frais").order_by("-date_paiement")
        ctx.update({"paiements":paiements,"total":paiements.aggregate(t=Sum("montant"))["t"] or 0})
        return render(request,"core/docs/recu_annuel.html",ctx)

    from django.contrib import messages as dj_messages
    dj_messages.error(request,"Type de document inconnu.")
    return redirect("liste_documents")

@login_required
@req
def carte_scolaire(request, eleve_pk):
    etab=request.etablissement
    eleve=get_object_or_404(Eleve,pk=eleve_pk,etablissement=etab)
    annee=AnneeScolaire.objects.filter(etablissement=etab,is_active=True).first()
    inscription=eleve.get_inscription_active()
    modele=get_modele(etab,'carte_scolaire')
    return render(request,"core/docs/carte_scolaire.html",{
        "eleve":eleve,"etab":etab,"annee":annee,
        "inscription":inscription,"modele":modele,"today":timezone.now().date(),
    })

@login_required
@req
def export_eleves_excel(request):
    import openpyxl; from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    etab=request.etablissement; annee=AnneeScolaire.objects.filter(etablissement=etab,is_active=True).first()
    eleves=get_eleves_actifs(etab).select_related("tuteur").prefetch_related("inscriptions__classe__niveau").order_by("nom","prenom")
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Eleves"
    hf=PatternFill(start_color="1565C0",end_color="1565C0",fill_type="solid"); hft=Font(color="FFFFFF",bold=True)
    b=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    ws.merge_cells("A1:H1"); ws["A1"]=f"LISTE DES ELEVES — {etab.nom}"; ws["A1"].font=Font(bold=True,size=14)
    headers=["#","Matricule","Nom","Prenom","Sexe","Naissance","Classe","Tuteur/Tel"]
    for c,h in enumerate(headers,1):
        cell=ws.cell(row=3,column=c,value=h); cell.fill=hf; cell.font=hft; cell.border=b
    af=PatternFill(start_color="EBF2FF",end_color="EBF2FF",fill_type="solid")
    for i,e in enumerate(eleves,1):
        insc=e.get_inscription_active(); f=af if i%2==0 else None
        data=[i,e.matricule,e.nom,e.prenom,e.get_sexe_display(),e.date_naissance.strftime("%d/%m/%Y"),insc.classe.nom if insc else "—",f"{e.tuteur.nom} {e.tuteur.prenom} — {e.tuteur.telephone}" if e.tuteur else "—"]
        for c,v in enumerate(data,1):
            cell=ws.cell(row=i+3,column=c,value=v); cell.border=b
            if f: cell.fill=f
    for c,w in enumerate([5,16,18,18,10,14,10,35],1): ws.column_dimensions[ws.cell(row=3,column=c).column_letter].width=w
    resp=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"]=f"attachment; filename=eleves_{etab.code}.xlsx"; wb.save(resp); return resp

@login_required
@req
def export_paiements_excel(request):
    import openpyxl; from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    etab=request.etablissement
    paiements=Paiement.objects.filter(etablissement=etab,statut="valide").select_related("eleve","type_frais").order_by("-date_paiement")
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Paiements"
    hf=PatternFill(start_color="2E7D32",end_color="2E7D32",fill_type="solid"); hft=Font(color="FFFFFF",bold=True)
    b=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    ws.merge_cells("A1:G1"); ws["A1"]=f"PAIEMENTS — {etab.nom}"; ws["A1"].font=Font(bold=True,size=14)
    for c,h in enumerate(["Reference","Eleve","Matricule","Type","Montant","Mode","Date"],1):
        cell=ws.cell(row=3,column=c,value=h); cell.fill=hf; cell.font=hft; cell.border=b
    total=0; af=PatternFill(start_color="E8F5E9",end_color="E8F5E9",fill_type="solid")
    pl=list(paiements)
    for i,p in enumerate(pl,1):
        f=af if i%2==0 else None
        data=[p.reference,f"{p.eleve.nom} {p.eleve.prenom}",p.eleve.matricule,p.type_frais.nom,float(p.montant),p.get_mode_paiement_display(),p.date_paiement.strftime("%d/%m/%Y")]
        for c,v in enumerate(data,1):
            cell=ws.cell(row=i+3,column=c,value=v); cell.border=b
            if f: cell.fill=f
            if c==5: cell.number_format="#,##0"
        total+=float(p.montant)
    last=len(pl)+4
    ws.cell(row=last,column=4,value="TOTAL").font=Font(bold=True)
    tc=ws.cell(row=last,column=5,value=total); tc.font=Font(bold=True); tc.number_format="#,##0"
    resp=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"]=f"attachment; filename=paiements_{etab.code}.xlsx"; wb.save(resp); return resp
